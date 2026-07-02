#!/usr/bin/env python3
# Convert the client's "Matrix - Enquiry (Durian).xlsx" sheet into the
# crm_owner_routing YAML blocks for routing_rules.yaml.
#
# Columns used (sheet "Zoho CRM - Enquiry", data from row 3):
#   E  Select Your Location Type → the location key the AI matches against
#   F  CRM Email Id              → owner_email (audit display)
#   J  CRM ID                    → owner_id (the Zoho user the record is
#                                  assigned to — col K is ignored: it's a
#                                  partial duplicate with 18 #N/A holes)
#   O  Business Vertical         → Furniture / Doors (blank → Furniture)
#   P  CRM Layout                → all "Standard" = Zoho's default layout, so
#                                  nothing to set at create time
#   C  Store Type (COCO/FOFO)    → IGNORED (per client, for now)
#   S  Keyword Bucket            → IGNORED (pincode mapping = phase 2)
#
# Rows with Business Vertical = Doors become crm_owner_routing_doors
# (bangalore / other), mirroring the doors email routing. Rows missing a CRM
# ID are skipped and reported. Run:
#   venv/bin/python convert_crm_owner_matrix.py "<path to xlsx>"

import re
import sys

import openpyxl


def clean(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def main(path: str):
    ws = openpyxl.load_workbook(path, data_only=True)["Zoho CRM - Enquiry"]
    furniture, doors, skipped = {}, {}, []
    for r in range(3, ws.max_row + 1):
        loc = clean(ws.cell(row=r, column=5).value)
        if not loc:
            continue
        email = clean(ws.cell(row=r, column=6).value)
        oid   = clean(ws.cell(row=r, column=10).value)
        vert  = clean(ws.cell(row=r, column=15).value) or "Furniture"
        if not re.fullmatch(r"\d{15,25}", oid):
            skipped.append((r, loc, "missing/invalid CRM ID (col J)"))
            continue
        entry = {"owner_email": email, "owner_id": oid,
                 "business_vertical": vert}
        if vert.lower() == "doors":
            key = "bangalore" if loc.lower() == "bangalore" else "other"
            doors[key] = entry
        else:
            if loc in furniture:
                skipped.append((r, loc, "duplicate location key"))
                continue
            furniture[loc] = entry

    print("# ── CRM owner routing (generated from the client's matrix sheet) ──")
    print("# Regenerate with convert_crm_owner_matrix.py when the sheet changes.")
    print("crm_owner_routing:")
    for loc, e in furniture.items():
        print(f"  {loc!r}:")
        print(f"    owner_email: {e['owner_email']}")
        print(f"    owner_id: \"{e['owner_id']}\"")
        print(f"    business_vertical: {e['business_vertical']}")
    print("  # Govt / CPWD deals — owner NOT in the client sheet; fill in the")
    print("  # Delhi-office/govt owner before enabling government deals:")
    print("  # govt:")
    print("  #   owner_email: <govt owner email>")
    print("  #   owner_id: \"<govt owner Zoho user id>\"")
    print("  #   business_vertical: Furniture")
    print()
    print("# Doors enquiries route Bangalore vs everywhere-else (sheet rows with")
    print("# Business Vertical = Doors), mirroring the doors email routing.")
    print("crm_owner_routing_doors:")
    for key, e in doors.items():
        print(f"  {key}:")
        print(f"    owner_email: {e['owner_email']}")
        print(f"    owner_id: \"{e['owner_id']}\"")
        print(f"    business_vertical: {e['business_vertical']}")

    print(f"\n# {len(furniture)} furniture locations, {len(doors)} doors desks.",
          file=sys.stderr)
    for r, loc, why in skipped:
        print(f"# SKIPPED row {r}: {loc!r} — {why}", file=sys.stderr)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else
         "/Users/siddharthsingh/Downloads/Matrix - Enquiry (Durian).xlsx")
