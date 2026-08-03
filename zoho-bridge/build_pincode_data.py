#!/usr/bin/env python3
# OFFLINE build tool (dev-only) — regenerates the bundled data the runtime
# pincode_resolver.py reads. Run occasionally when the client updates the
# pincode sheet or the store list; the JSON it writes is committed to the repo
# so the bridge has ZERO runtime geocoding dependency.
#
# Needs: openpyxl + pgeocode (NOT bridge runtime deps — install ad hoc):
#   python -m pip install openpyxl pgeocode
#   python build_pincode_data.py \
#       "Pincode Final for CRM.xlsx" "City wise Location template - Checked.xlsx"
#
# Produces (in ./data):
#   pincode_geo.json   {pincode: [lat, lon]}      — the offline geocoder
#   pincode_tags.json  {tags:[...], pins:{pin:i}} — furniture pincode → showroom
#   nonfurniture_stores.json  [{vertical, store, city, lat, lon}]  — doors/FHC
#
# Design: furniture routing is the client's own per-pincode assignment
# ("Tagged Showroom for Lead"); doors/FHC have no such tagging, so those stores
# are geocoded (by a representative city pincode) for nearest-by-distance.

import json
import sys
from pathlib import Path

import openpyxl
import pgeocode

_OUT = Path(__file__).parent / "data"

# Curated city → (lat, lon) for the doors/FHC store cities. These few stores are
# widely separated, so city-level coords pick the nearest reliably — and a hand-
# checked table avoids fuzzy-geocoding mistakes (e.g. two "Goregaon"s in India).
# Extend when a new doors/FHC city opens.
_CITY_COORDS = {
    "delhi ncr": (28.61, 77.21), "delhi": (28.61, 77.21),
    "mumbai": (19.16, 72.85), "goregaon": (19.16, 72.85),  # Goregaon = Mumbai
    "bangalore": (12.97, 77.59), "bengaluru": (12.97, 77.59),
    "gurgaon": (28.46, 77.03), "gurugram": (28.46, 77.03),
    "coimbatore": (11.02, 76.97), "hyderabad": (17.39, 78.49),
    "indore": (22.72, 75.86), "lucknow": (26.85, 80.95),
    "bikaner": (28.02, 73.31),
}


def _pin6(v) -> str | None:
    try:
        return str(int(float(v))).zfill(6)
    except (TypeError, ValueError):
        return None


def _col(ws, name: str) -> int | None:
    hdr = [str(ws.cell(1, c).value or "").strip().lower()
           for c in range(1, ws.max_column + 1)]
    for i, h in enumerate(hdr, 1):
        if h == name.lower() or h.startswith(name.lower()):
            return i
    return None


def build(pincode_xlsx: str, stores_xlsx: str) -> None:
    _OUT.mkdir(exist_ok=True)
    nomi = pgeocode.Nominatim("in")

    # 1. Full offline geocoder: every India pincode pgeocode knows → [lat, lon].
    geo = {}
    df = nomi._data[["postal_code", "latitude", "longitude"]].dropna()
    for pc, la, lo in df.itertuples(index=False):
        p = _pin6(pc)
        if p and p not in geo:
            geo[p] = [round(float(la), 4), round(float(lo), 4)]
    (_OUT / "pincode_geo.json").write_text(json.dumps(geo, separators=(",", ":")))
    print(f"pincode_geo.json: {len(geo)} pincodes")

    # 2. Furniture pincode → showroom tag (client's own assignment). Interned.
    wb = openpyxl.load_workbook(pincode_xlsx, data_only=True)
    ws = wb["All India Pincode"]
    cPin, cTag = _col(ws, "Pincode"), _col(ws, "Tagged Showroom for Lead")
    tag_list, tag_idx, pins = [], {}, {}
    for r in range(2, ws.max_row + 1):
        p = _pin6(ws.cell(r, cPin).value)
        t = ws.cell(r, cTag).value
        if not p or t in (None, "", "#N/A"):
            continue
        t = str(t).strip()
        if t not in tag_idx:
            tag_idx[t] = len(tag_list)
            tag_list.append(t)
        pins[p] = tag_idx[t]
    (_OUT / "pincode_tags.json").write_text(
        json.dumps({"tags": tag_list, "pins": pins}, separators=(",", ":")))
    print(f"pincode_tags.json: {len(pins)} tagged pincodes, {len(tag_list)} showrooms")

    # 3. Doors / FHC stores → coords from the curated city table.
    wb2 = openpyxl.load_workbook(stores_xlsx, data_only=True)
    ws2 = next(wb2[s] for s in wb2.sheetnames if s.strip().lower().startswith("stores details"))
    cV, cCity, cName = _col(ws2, "Vertical"), _col(ws2, "City"), _col(ws2, "Showroom Name")
    stores, missing = [], set()
    for r in range(2, ws2.max_row + 1):
        v = str(ws2.cell(r, cV).value or "").strip().lower()
        vert = "doors" if v == "doors" else "fhc" if v == "full home customisation" else None
        if not vert:
            continue
        city = str(ws2.cell(r, cCity).value or "").strip()
        coords = _CITY_COORDS.get(city.lower())
        if not coords:
            missing.add(city)
            continue
        stores.append({"vertical": vert, "store": str(ws2.cell(r, cName).value or "").strip(),
                       "city": city, "lat": coords[0], "lon": coords[1]})
    (_OUT / "nonfurniture_stores.json").write_text(json.dumps(stores, indent=1))
    print(f"nonfurniture_stores.json: {len(stores)} doors/FHC stores")
    if missing:
        print(f"  WARNING: no city pincode for {sorted(missing)} — add to _CITY_PIN")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit(__doc__)
    build(sys.argv[1], sys.argv[2])
