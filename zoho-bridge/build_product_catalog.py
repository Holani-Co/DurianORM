#!/usr/bin/env python3
# OFFLINE build tool (dev-only) — builds the product catalog the EMI / product
# flows read, from the client's monthly price workbook. Run when the client
# sends a new month's price list; the JSON it writes is committed to the repo so
# the bridge has no spreadsheet dependency at runtime.
#
#   python -m pip install openpyxl
#   python build_product_catalog.py "REVISED SALE PRICE ... .xlsx" [price_period]
#
# Reads two tabs (verified schema, header on row 2, data from row 3):
#   'Durian' : col1 SKU ID, col2 MRP, col3 BP, col4 Offer price (SALE),
#              col9 CATEGORY
#   'RT'     : col1 No.(SKU), col2 Description (name), col4 MRP,
#              col8 BP Incl GST (REVISED SALE price), col10 Category
#
# The customer-facing SALE price (= MRP*(1-disc)) is Durian.Offer price; where a
# SKU is only in RT we fall back to RT's revised sale price (col8). Both are the
# same quantity — the build cross-checks and reports any mismatch.
#
# Output: data/product_catalog.json
#   { "price_period": "...", "products": {
#       "RIHANNA/A/2": {"name","family","mrp","sale_price","category","source"} } }

import json
import sys
from pathlib import Path

import openpyxl

_OUT = Path(__file__).parent / "data" / "product_catalog.json"


def _num(v):
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _sku(v):
    s = str(v or "").strip().upper()
    return s if s and s != "#N/A" else None


def _clean(v):
    s = str(v or "").strip()
    return "" if s in ("", "#N/A", "None") else s


def build(xlsx: str, price_period: str = "") -> None:
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # 1. Durian tab → the sale-price authority.
    durian = {}
    ws = wb["Durian"]
    for r in range(3, (ws.max_row or 0) + 1):
        sku = _sku(ws.cell(r, 1).value)
        if not sku:
            continue
        durian[sku] = {"mrp": _num(ws.cell(r, 2).value),
                       "offer": _num(ws.cell(r, 4).value),
                       "category": _clean(ws.cell(r, 9).value)}

    # 2. RT tab → SKU → description (name) + revised sale price.
    rt = {}
    ws = wb["RT"]
    for r in range(3, (ws.max_row or 0) + 1):
        sku = _sku(ws.cell(r, 1).value)
        if not sku:
            continue
        rt[sku] = {"name": _clean(ws.cell(r, 2).value),
                   "mrp": _num(ws.cell(r, 4).value),
                   "revised_sale": _num(ws.cell(r, 8).value),
                   "category": _clean(ws.cell(r, 10).value)}

    # 3. Merge on SKU. sale_price prefers Durian.Offer, else RT.revised_sale.
    products, mismatches = {}, 0
    for sku in sorted(set(durian) | set(rt)):
        d, t = durian.get(sku, {}), rt.get(sku, {})
        offer, revised = d.get("offer"), t.get("revised_sale")
        # cross-check where both exist (>1% apart is worth flagging)
        if offer and revised and abs(offer - revised) / max(offer, revised) > 0.01:
            mismatches += 1
        sale = offer or revised
        if not sale:
            continue                      # no sale price → can't quote EMI/price
        products[sku] = {
            "name": t.get("name") or "",
            "family": sku.split("/")[0],
            "mrp": d.get("mrp") or t.get("mrp"),
            "sale_price": round(sale),
            "category": d.get("category") or t.get("category") or "",
            "source": "durian" if offer else "rt",
        }

    _OUT.parent.mkdir(exist_ok=True)
    _OUT.write_text(json.dumps({"price_period": price_period, "products": products},
                               separators=(",", ":")))
    named = sum(1 for p in products.values() if p["name"])
    from_rt = sum(1 for p in products.values() if p["source"] == "rt")
    print(f"product_catalog.json: {len(products)} products "
          f"({named} with a name, {from_rt} priced from RT fallback)")
    print(f"  cross-check: {mismatches} SKUs where Durian.Offer vs RT.revised differ >1%")
    print(f"  Durian rows: {len(durian)} | RT rows: {len(rt)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    build(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "")
