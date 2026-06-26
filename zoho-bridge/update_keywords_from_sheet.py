#!/usr/bin/env python3
# One-time: merge the client's "Email - Keywords" spreadsheet into the
# `keywords:` lists of routing_rules.yaml — UNION with what's already there,
# deduped case-insensitively, order preserved (existing first, then new).
# Only the keywords lists are rewritten; descriptions, examples, team_ids,
# comments, and structure are left untouched.
#
#   python update_keywords_from_sheet.py "/path/to/Feedback.xlsx"
#
# Re-runnable: re-importing the same sheet adds nothing (already merged).

import re
import sys

import openpyxl
import yaml

YAML_PATH = "routing_rules.yaml"

# Spreadsheet column (1-indexed) → routing category key.
COL_TO_CAT = {
    1: "career_job_enquiry",
    2: "complaint",
    3: "legal_complaint",
    4: "collaboration_request",
    5: "marketing_advertising",
    6: "vendor_supplier_enquiry",
    7: "franchise_dealership",
    8: "project_bulk_order",
    9: "full_home_customization",
    10: "doors_veneer_plywood",
}


def _norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def _sheet_keywords(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Email - Keywords"]
    out = {}
    for col, cat in COL_TO_CAT.items():
        seen, kws = set(), []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v and _norm(v) and _norm(v).lower() not in seen:
                seen.add(_norm(v).lower())
                kws.append(_norm(v))
        out[cat] = kws
    return out


def main():
    if len(sys.argv) < 2:
        raise SystemExit("usage: python update_keywords_from_sheet.py <xlsx>")
    sheet = _sheet_keywords(sys.argv[1])
    rules = yaml.safe_load(open(YAML_PATH))
    cats = rules["categories"]

    # Compute the merged (current ∪ sheet) list per category.
    merged = {}
    for cat, new_kws in sheet.items():
        cur = [_norm(k) for k in (cats.get(cat, {}).get("keywords") or [])]
        cur_l = {k.lower() for k in cur}
        added = [k for k in new_kws if k.lower() not in cur_l]
        merged[cat] = (cur + added, len(added))

    # Rewrite only the keywords blocks, in the original file text.
    lines = open(YAML_PATH).read().split("\n")
    out, i, cur_cat = [], 0, None
    cat_hdr = re.compile(r"^  ([a-z_]+):\s*$")
    while i < len(lines):
        line = lines[i]
        m = cat_hdr.match(line)
        if m and m.group(1) in cats:
            cur_cat = m.group(1)
            out.append(line)
            i += 1
            continue
        if cur_cat in merged and re.match(r"^    keywords:\s*$", line):
            out.append(line)
            i += 1
            # skip the existing "      - ..." items
            while i < len(lines) and re.match(r"^      - ", lines[i]):
                i += 1
            # write the merged list
            for kw in merged[cur_cat][0]:
                out.append(f"      - {kw}")
            continue
        out.append(line)
        i += 1

    new_text = "\n".join(out)
    # Validate before writing.
    parsed = yaml.safe_load(new_text)
    for cat, (full, added) in merged.items():
        got = parsed["categories"][cat].get("keywords") or []
        assert len(got) == len(full), f"{cat}: wrote {len(got)} expected {len(full)}"
    open(YAML_PATH, "w").write(new_text)

    print("Merged keywords into routing_rules.yaml:")
    for cat, (full, added) in merged.items():
        print(f"  {cat:28} +{added:>3} new  → {len(full)} total")


if __name__ == "__main__":
    main()
